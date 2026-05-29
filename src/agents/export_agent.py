import os
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from src.agents.base import BaseAgent
from src.config import settings
from src.models import EventType, Lead, LeadStatus, OutputFormat, PipelineRun, Segment


COLUMNS = [
    "full_name",
    "first_name",
    "last_name",
    "title",
    "company",
    "company_domain",
    "email",
    "email_confidence",
    "phone",
    "location",
    "linkedin_url",
    "company_linkedin_url",
    "intent_score",
    "segment",
    "status",
]


class ExportAgent(BaseAgent):
    def __init__(self, run: PipelineRun, leads: list[Lead]):
        super().__init__(run)
        self.leads = leads

    def _to_df(self, leads: list[Lead]) -> pd.DataFrame:
        rows = []
        for lead in leads:
            data = lead.to_dict()
            data["segment"] = lead.segment.value
            data["status"] = lead.status.value
            rows.append({column: data.get(column, "") for column in COLUMNS})
        return pd.DataFrame(rows, columns=COLUMNS)

    def _style_sheet(self, ws, hex_color: str) -> None:
        fill = PatternFill(
            start_color=hex_color,
            end_color=hex_color,
            fill_type="solid",
        )
        for cell in ws[1]:
            cell.fill = fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        for column in ws.columns:
            width = max((len(str(cell.value or "")) for cell in column), default=8)
            ws.column_dimensions[column[0].column_letter].width = min(width + 4, 50)

    def _export_xlsx(self, timestamp: str) -> list[str]:
        path = settings.output_dir / f"leads_{timestamp}.xlsx"
        warm = [lead for lead in self.leads if lead.segment == Segment.WARM]
        cold = [lead for lead in self.leads if lead.segment == Segment.COLD]
        no_email = [lead for lead in self.leads if lead.segment == Segment.NO_EMAIL]
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            for name, data, color in [
                ("Warm", warm, "4CAF50"),
                ("Cold", cold, "2196F3"),
                ("No_Email", no_email, "9E9E9E"),
            ]:
                self._to_df(data).to_excel(writer, sheet_name=name, index=False)
                self._style_sheet(writer.sheets[name], color)
        return [str(path)]

    def _export_csv(self, timestamp: str) -> list[str]:
        paths = []
        for segment, data in [
            ("warm", [lead for lead in self.leads if lead.segment == Segment.WARM]),
            ("cold", [lead for lead in self.leads if lead.segment == Segment.COLD]),
            (
                "no_email",
                [lead for lead in self.leads if lead.segment == Segment.NO_EMAIL],
            ),
        ]:
            path = settings.output_dir / f"leads_{segment}_{timestamp}.csv"
            self._to_df(data).to_csv(path, index=False, encoding="utf-8-sig")
            paths.append(str(path))
        return paths

    def run_agent(self) -> list[str]:
        from datetime import datetime

        settings.output_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        paths = (
            self._export_xlsx(timestamp)
            if settings.output_format == OutputFormat.XLSX
            else self._export_csv(timestamp)
        )
        self.run.total_exported = len(self.leads)
        for path in paths:
            self.emit(EventType.LEAD_EXPORTED, {"file": path})
            for lead in self.leads:
                lead.status = LeadStatus.EXPORTED
        return paths
